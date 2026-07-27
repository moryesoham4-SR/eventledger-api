"""
Whole-event export/import: one .xlsx workbook covering everything on an
event — Departments, Budget Proposals + line items, Estimated/Actual
Expenses, Estimated/Actual Income, Vendors, and Sponsors — as separate
sheets, so someone can back up, edit offline, or clone the financial setup
of an event in one file instead of doing it screen by screen.

Restricted to event_admin/finance_head (same bar as budget import) since
it can create departments and touches every financial table on the event.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from core.database import get_db, execute
from core.auth import get_current_user
from utils.roles import get_event_role
import io
import openpyxl
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/events", tags=["event-data"])

SHEETS = {
    "Departments": ["Name", "Head Name", "Color"],
    "Budget Proposals": [
        "Department", "Proposal Title", "Status", "Category", "Item Name",
        "Description", "Quantity", "Unit", "Unit Price", "Total Amount",
    ],
    "Estimated Expenses": [
        "Department", "Category", "Item Name", "Description", "Quantity", "Unit", "Amount", "Notes",
    ],
    "Actual Expenses": [
        "Department", "Category", "Item Name", "Description", "Quantity", "Unit", "Amount",
        "Paid On", "Payment Mode", "Status", "Reference", "Notes",
    ],
    "Estimated Income": ["Source", "Category", "Amount", "Notes"],
    "Actual Income": ["Source", "Category", "Amount", "Received On", "Payment Mode", "Reference", "Notes"],
    "Vendors": ["Name", "Category", "Contact Name", "Contact Email", "Contract Value", "Status", "Notes"],
    "Sponsors": ["Name", "Tier", "Contact Name", "Contact Email", "Amount", "Status", "Notes"],
}

# --- Import robustness: recognize a sheet/column by MEANING, not exact text ---
# People rename sheets and reorder/rename columns when they edit a spreadsheet by
# hand — someone uploading "just an expenses file" shouldn't have it silently
# ignored because their sheet is called "Expenses" instead of "Actual Expenses",
# or their columns are in a different order than our export.
import re

def _norm(s):
    """'Contact  E-mail' -> 'contactemail' — lets us match headers loosely."""
    return re.sub(r'[^a-z0-9]', '', str(s).lower()) if s not in (None, '') else ''

SHEET_ALIASES = {
    "Departments": {"departments", "department", "depts", "dept"},
    "Budget Proposals": {"budgetproposals", "budgetproposal", "budget", "proposals", "proposal"},
    "Estimated Expenses": {"estimatedexpenses", "estexpenses", "expensesestimated", "estimatedexpense", "budgetedexpenses"},
    "Actual Expenses": {"actualexpenses", "actualexpense", "expenses", "expense"},
    "Estimated Income": {"estimatedincome", "estincome", "incomeestimated", "estimatedincomes", "budgetedincome"},
    "Actual Income": {"actualincome", "actualincomes", "income", "incomes"},
    "Vendors": {"vendors", "vendor"},
    "Sponsors": {"sponsors", "sponsor"},
}

# canonical field name -> acceptable normalized header text within that field's sheet
FIELD_ALIASES = {
    "Department": {"department", "dept", "departmentname"},
    "Head Name": {"headname", "head", "departmenthead"},
    "Color": {"color", "colour"},
    "Proposal Title": {"proposaltitle", "title", "proposal", "name"},
    "Category": {"category", "cat", "type"},
    "Item Name": {"itemname", "item", "name", "particulars"},
    "Description": {"description", "desc", "remarks", "remark"},
    "Quantity": {"quantity", "qty", "units"},
    "Unit": {"unit", "uom"},
    "Unit Price": {"unitprice", "price", "rate", "unitcost"},
    "Total Amount": {"totalamount", "total"},
    "Amount": {"amount", "value", "amt", "cost", "expense", "price", "total"},
    "Notes": {"notes", "note", "comment", "comments", "remarks", "remark"},
    "Paid On": {"paidon", "paiddate", "date", "expensedate"},
    "Payment Mode": {"paymentmode", "mode", "paymentmethod", "paidvia"},
    "Status": {"status"},
    "Reference": {"reference", "ref", "refno", "referenceno", "receiptno"},
    "Source": {"source", "incomesource", "name"},
    "Received On": {"receivedon", "receiveddate", "date"},
    "Name": {"name", "vendorname", "sponsorname"},
    "Contact Name": {"contactname", "contact", "contactperson"},
    "Contact Email": {"contactemail", "email"},
    "Contract Value": {"contractvalue", "value", "contract"},
    "Tier": {"tier", "level", "category"},
}


def _require_admin_or_finance(conn, user, event_id):
    role_ctx = get_event_role(conn, user, event_id)
    if role_ctx["level"] not in ("event_admin", "finance_head"):
        raise HTTPException(status_code=403, detail="Only an event admin or finance head can do this")
    return role_ctx


def _write_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 2)


@router.get("/{event_id}/export")
def export_event(event_id: int, conn=Depends(get_db), user=Depends(get_current_user)):
    _require_admin_or_finance(conn, user, event_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    cur = execute(conn, "SELECT name, head_name, color FROM departments WHERE event_id=%s ORDER BY name", (event_id,))
    _write_sheet(wb, "Departments", SHEETS["Departments"], [[r["name"], r["head_name"], r["color"]] for r in cur.fetchall()])

    cur = execute(conn, """
        SELECT d.name AS dept_name, p.title, p.status, li.category, li.item_name,
               li.description, li.quantity, li.unit, li.unit_price, li.total_amount
        FROM budget_proposals p
        JOIN departments d ON d.id = p.department_id
        LEFT JOIN budget_line_items li ON li.proposal_id = p.id
        WHERE p.event_id=%s ORDER BY d.name, p.title, li.id
    """, (event_id,))
    _write_sheet(wb, "Budget Proposals", SHEETS["Budget Proposals"], [
        [r["dept_name"], r["title"], r["status"], r["category"], r["item_name"],
         r["description"], r["quantity"], r["unit"], r["unit_price"], r["total_amount"]]
        for r in cur.fetchall()
    ])

    cur = execute(conn, """
        SELECT d.name AS dept_name, e.category, e.item_name, e.description, e.quantity, e.unit, e.amount, e.notes
        FROM estimated_expenses e LEFT JOIN departments d ON d.id = e.department_id
        WHERE e.event_id=%s ORDER BY e.created_at
    """, (event_id,))
    _write_sheet(wb, "Estimated Expenses", SHEETS["Estimated Expenses"], [
        [r["dept_name"], r["category"], r["item_name"], r["description"], r["quantity"], r["unit"], r["amount"], r["notes"]]
        for r in cur.fetchall()
    ])

    cur = execute(conn, """
        SELECT d.name AS dept_name, e.category, e.item_name, e.description, e.quantity, e.unit, e.amount,
               e.paid_on, e.payment_mode, e.status, e.reference, e.notes
        FROM actual_expenses e LEFT JOIN departments d ON d.id = e.department_id
        WHERE e.event_id=%s ORDER BY e.created_at
    """, (event_id,))
    _write_sheet(wb, "Actual Expenses", SHEETS["Actual Expenses"], [
        [r["dept_name"], r["category"], r["item_name"], r["description"], r["quantity"], r["unit"], r["amount"],
         r["paid_on"], r["payment_mode"], r["status"], r["reference"], r["notes"]]
        for r in cur.fetchall()
    ])

    cur = execute(conn, "SELECT source, category, amount, notes FROM estimated_income WHERE event_id=%s ORDER BY created_at", (event_id,))
    _write_sheet(wb, "Estimated Income", SHEETS["Estimated Income"], [[r["source"], r["category"], r["amount"], r["notes"]] for r in cur.fetchall()])

    cur = execute(conn, """
        SELECT source, category, amount, received_on, payment_mode, reference, notes
        FROM actual_income WHERE event_id=%s ORDER BY created_at
    """, (event_id,))
    _write_sheet(wb, "Actual Income", SHEETS["Actual Income"], [
        [r["source"], r["category"], r["amount"], r["received_on"], r["payment_mode"], r["reference"], r["notes"]]
        for r in cur.fetchall()
    ])

    cur = execute(conn, """
        SELECT name, category, contact_name, contact_email, contract_value, status, notes
        FROM vendors WHERE event_id=%s ORDER BY created_at
    """, (event_id,))
    _write_sheet(wb, "Vendors", SHEETS["Vendors"], [
        [r["name"], r["category"], r["contact_name"], r["contact_email"], r["contract_value"], r["status"], r["notes"]]
        for r in cur.fetchall()
    ])

    cur = execute(conn, """
        SELECT name, tier, contact_name, contact_email, amount, status, notes
        FROM sponsors WHERE event_id=%s ORDER BY created_at
    """, (event_id,))
    _write_sheet(wb, "Sponsors", SHEETS["Sponsors"], [
        [r["name"], r["tier"], r["contact_name"], r["contact_email"], r["amount"], r["status"], r["notes"]]
        for r in cur.fetchall()
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=event_export_{event_id}.xlsx"},
    )


@router.post("/{event_id}/import")
async def import_event(event_id: int, file: UploadFile = File(...), conn=Depends(get_db), user=Depends(get_current_user)):
    _require_admin_or_finance(conn, user, event_id)

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Couldn't read that file — is it a valid .xlsx?")

    matched_sheet_names = set()

    def read_sheet(canonical_name):
        """Finds the sheet matching `canonical_name` by name or alias, then
        returns (rows, header_map) where header_map maps normalized header
        text -> column index, so fields can be pulled by meaning rather than
        position. Returns ([], {}) if nothing in the file matches."""
        target_aliases = SHEET_ALIASES.get(canonical_name, set()) | {_norm(canonical_name)}
        actual_name = next((sn for sn in wb.sheetnames if _norm(sn) in target_aliases), None)
        if not actual_name:
            return [], {}
        matched_sheet_names.add(actual_name)
        ws = wb[actual_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return [], {}
        header_map = {_norm(h): i for i, h in enumerate(header_row) if h not in (None, '')}
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and any(r)]
        return rows, header_map

    def field(row, header_map, canonical_field):
        for alias in FIELD_ALIASES.get(canonical_field, {_norm(canonical_field)}):
            idx = header_map.get(alias)
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    summary = {"departments_created": 0, "budget_proposals_created": 0, "budget_items_created": 0,
               "expenses_created": 0, "income_created": 0, "vendors_created": 0, "sponsors_created": 0}
    errors = []

    # 1. Departments first — everything else may reference a department by name.
    cur = execute(conn, "SELECT id, name FROM departments WHERE event_id=%s", (event_id,))
    dept_map = {row["name"].strip().lower(): row["id"] for row in cur.fetchall()}

    dept_rows, dept_headers = read_sheet("Departments")
    for row in dept_rows:
        name = field(row, dept_headers, "Name")
        head_name = field(row, dept_headers, "Head Name")
        color = field(row, dept_headers, "Color")
        if not name:
            continue
        key = str(name).strip().lower()
        if key in dept_map:
            continue
        cur = execute(conn, "INSERT INTO departments (event_id,name,head_name,color) VALUES (%s,%s,%s,%s) RETURNING id",
                      (event_id, str(name).strip(), head_name or "", color or "#6366f1"))
        dept_map[key] = cur.fetchone()["id"]
        summary["departments_created"] += 1

    def dept_id_for(name, row_label):
        if not name:
            return None, True
        did = dept_map.get(str(name).strip().lower())
        if not did:
            errors.append(f"{row_label}: no department named '{name}'")
            return None, False
        return did, True

    # 2. Budget proposals + line items
    proposal_cache = {}
    bp_rows, bp_headers = read_sheet("Budget Proposals")
    for idx, row in enumerate(bp_rows, start=2):
        dept_name = field(row, bp_headers, "Department")
        title = field(row, bp_headers, "Proposal Title")
        category = field(row, bp_headers, "Category")
        item_name = field(row, bp_headers, "Item Name")
        description = field(row, bp_headers, "Description")
        quantity = field(row, bp_headers, "Quantity")
        unit = field(row, bp_headers, "Unit")
        unit_price = field(row, bp_headers, "Unit Price")
        total_amount = field(row, bp_headers, "Total Amount")

        if not dept_name or not title:
            errors.append(f"Budget Proposals row {idx}: missing department or title")
            continue
        did, ok = dept_id_for(dept_name, f"Budget Proposals row {idx}")
        if not ok:
            continue
        key = (did, str(title).strip())
        if key not in proposal_cache:
            cur = execute(conn,
                "INSERT INTO budget_proposals (event_id,department_id,submitted_by,title,notes,status) VALUES (%s,%s,%s,%s,'','draft') RETURNING id",
                (event_id, did, user["id"], str(title).strip()))
            proposal_cache[key] = cur.fetchone()["id"]
            summary["budget_proposals_created"] += 1
        proposal_id = proposal_cache[key]
        if category or item_name:
            try:
                qty = float(quantity) if quantity not in (None, "") else 1
                price = float(unit_price) if unit_price not in (None, "") else 0
            except (TypeError, ValueError):
                errors.append(f"Budget Proposals row {idx}: quantity/unit price must be numbers")
                continue
            total = float(total_amount) if total_amount not in (None, "") else qty * price
            execute(conn,
                """INSERT INTO budget_line_items (proposal_id,category,item_name,description,quantity,unit,unit_price,total_amount)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                (proposal_id, category or "", item_name or "", description or "", qty, unit or "unit", price, total))
            summary["budget_items_created"] += 1

    # 3. Estimated expenses
    ee_rows, ee_headers = read_sheet("Estimated Expenses")
    for idx, row in enumerate(ee_rows, start=2):
        dept_name = field(row, ee_headers, "Department")
        category = field(row, ee_headers, "Category")
        item_name = field(row, ee_headers, "Item Name")
        description = field(row, ee_headers, "Description")
        quantity = field(row, ee_headers, "Quantity")
        unit = field(row, ee_headers, "Unit")
        amount = field(row, ee_headers, "Amount")
        notes = field(row, ee_headers, "Notes")

        if not category or amount in (None, ""):
            errors.append(f"Estimated Expenses row {idx}: missing category or amount")
            continue
        did, ok = dept_id_for(dept_name, f"Estimated Expenses row {idx}")
        if not ok:
            continue
        try:
            qty = float(quantity) if quantity not in (None, "") else 1
            amt = float(amount)
        except (TypeError, ValueError):
            errors.append(f"Estimated Expenses row {idx}: quantity/amount must be numbers")
            continue
        execute(conn,
            """INSERT INTO estimated_expenses (event_id,department_id,category,item_name,description,quantity,unit,amount,notes)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (event_id, did, category, item_name or "", description or "", qty, unit or "unit", amt, notes or ""))
        summary["expenses_created"] += 1

    # 4. Actual expenses
    ae_rows, ae_headers = read_sheet("Actual Expenses")
    for idx, row in enumerate(ae_rows, start=2):
        dept_name = field(row, ae_headers, "Department")
        category = field(row, ae_headers, "Category")
        item_name = field(row, ae_headers, "Item Name")
        description = field(row, ae_headers, "Description")
        quantity = field(row, ae_headers, "Quantity")
        unit = field(row, ae_headers, "Unit")
        amount = field(row, ae_headers, "Amount")
        paid_on = field(row, ae_headers, "Paid On")
        payment_mode = field(row, ae_headers, "Payment Mode")
        status = field(row, ae_headers, "Status")
        reference = field(row, ae_headers, "Reference")
        notes = field(row, ae_headers, "Notes")

        if not category or amount in (None, ""):
            errors.append(f"Actual Expenses row {idx}: missing category or amount")
            continue
        did, ok = dept_id_for(dept_name, f"Actual Expenses row {idx}")
        if not ok:
            continue
        try:
            qty = float(quantity) if quantity not in (None, "") else 1
            amt = float(amount)
        except (TypeError, ValueError):
            errors.append(f"Actual Expenses row {idx}: quantity/amount must be numbers")
            continue
        execute(conn,
            """INSERT INTO actual_expenses (event_id,department_id,category,item_name,description,quantity,unit,amount,paid_on,payment_mode,status,reference,notes)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (event_id, did, category, item_name or "", description or "", qty, unit or "unit", amt,
             paid_on or "", payment_mode or "Cash", status or "paid", reference or "", notes or ""))
        summary["expenses_created"] += 1

    # 5. Estimated income
    ei_rows, ei_headers = read_sheet("Estimated Income")
    for idx, row in enumerate(ei_rows, start=2):
        source = field(row, ei_headers, "Source")
        category = field(row, ei_headers, "Category")
        amount = field(row, ei_headers, "Amount")
        notes = field(row, ei_headers, "Notes")

        if not source or amount in (None, ""):
            errors.append(f"Estimated Income row {idx}: missing source or amount")
            continue
        try:
            amt = float(amount)
        except (TypeError, ValueError):
            errors.append(f"Estimated Income row {idx}: amount must be a number")
            continue
        execute(conn, "INSERT INTO estimated_income (event_id,source,category,amount,notes) VALUES (%s,%s,%s,%s,%s)",
                (event_id, source, category or "Other", amt, notes or ""))
        summary["income_created"] += 1

    # 6. Actual income
    ai_rows, ai_headers = read_sheet("Actual Income")
    for idx, row in enumerate(ai_rows, start=2):
        source = field(row, ai_headers, "Source")
        category = field(row, ai_headers, "Category")
        amount = field(row, ai_headers, "Amount")
        received_on = field(row, ai_headers, "Received On")
        payment_mode = field(row, ai_headers, "Payment Mode")
        reference = field(row, ai_headers, "Reference")
        notes = field(row, ai_headers, "Notes")

        if not source or amount in (None, ""):
            errors.append(f"Actual Income row {idx}: missing source or amount")
            continue
        try:
            amt = float(amount)
        except (TypeError, ValueError):
            errors.append(f"Actual Income row {idx}: amount must be a number")
            continue
        execute(conn,
            "INSERT INTO actual_income (event_id,source,category,amount,received_on,payment_mode,reference,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (event_id, source, category or "Other", amt, received_on or "", payment_mode or "Cash", reference or "", notes or ""))
        summary["income_created"] += 1

    # 7. Vendors
    v_rows, v_headers = read_sheet("Vendors")
    for idx, row in enumerate(v_rows, start=2):
        name = field(row, v_headers, "Name")
        category = field(row, v_headers, "Category")
        contact_name = field(row, v_headers, "Contact Name")
        contact_email = field(row, v_headers, "Contact Email")
        contract_value = field(row, v_headers, "Contract Value")
        status = field(row, v_headers, "Status")
        notes = field(row, v_headers, "Notes")

        if not name:
            errors.append(f"Vendors row {idx}: missing name")
            continue
        try:
            cv = float(contract_value) if contract_value not in (None, "") else 0
        except (TypeError, ValueError):
            cv = 0
        execute(conn,
            "INSERT INTO vendors (event_id,name,category,contact_name,contact_email,contract_value,status,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (event_id, name, category or "Other", contact_name or "", contact_email or "", cv, status or "active", notes or ""))
        summary["vendors_created"] += 1

    # 8. Sponsors (also syncs to actual_income, same as the normal create-sponsor endpoint)
    s_rows, s_headers = read_sheet("Sponsors")
    for idx, row in enumerate(s_rows, start=2):
        name = field(row, s_headers, "Name")
        tier = field(row, s_headers, "Tier")
        contact_name = field(row, s_headers, "Contact Name")
        contact_email = field(row, s_headers, "Contact Email")
        amount = field(row, s_headers, "Amount")
        status = field(row, s_headers, "Status")
        notes = field(row, s_headers, "Notes")

        if not name:
            errors.append(f"Sponsors row {idx}: missing name")
            continue
        try:
            amt = float(amount) if amount not in (None, "") else 0
        except (TypeError, ValueError):
            amt = 0
        cur = execute(conn,
            "INSERT INTO sponsors (event_id,name,tier,contact_name,contact_email,amount,status,notes,income_synced) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,1) RETURNING id",
            (event_id, name, tier or "Bronze", contact_name or "", contact_email or "", amt, status or "confirmed", notes or ""))
        sponsor_id = cur.fetchone()["id"]
        execute(conn,
            "INSERT INTO actual_income (event_id,source,category,amount,payment_mode,sponsor_id) VALUES (%s,%s,'Sponsor',%s,'Bank Transfer',%s)",
            (event_id, f"Sponsor: {name}", amt, sponsor_id))
        summary["sponsors_created"] += 1

    # Sheets in the uploaded file that didn't match anything we recognize —
    # surfaced so a typo'd sheet name doesn't just silently import nothing.
    unrecognized = [sn for sn in wb.sheetnames if sn not in matched_sheet_names]
    if unrecognized:
        errors.append(f"Sheet(s) not recognized and skipped: {', '.join(unrecognized)}")

    return {"ok": True, **summary, "errors": errors}
